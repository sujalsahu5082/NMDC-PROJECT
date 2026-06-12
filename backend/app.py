"""
NMDC HR Analytics Intelligence — Flask Backend
================================================
REST API that handles Excel/CSV uploads, server-side data parsing,
normalization, and analytics computation.

Endpoints:
  GET  /                  → Serves the frontend index.html
  POST /api/upload        → Parse uploaded Excel/CSV file; returns headers + preview rows
  POST /api/process       → Apply column mapping to a previously uploaded file; stores normalized records
  GET  /api/analytics     → Returns aggregated analytics JSON for current filters
  GET  /api/summary       → Returns sidebar counts
  DELETE /api/clear       → Clears all server-side data
"""

import os
import io
import json
import uuid
import re
from pathlib import Path
from collections import defaultdict

import pandas as pd
from flask import Flask, request, jsonify, send_file, session
from flask_cors import CORS

# ─── App Setup ───────────────────────────────────────────────────────────────
BASE_DIR     = Path(__file__).parent          # .../NMDC DASHBOARD/backend
FRONTEND_DIR = BASE_DIR.parent / 'frontend'   # .../NMDC DASHBOARD/frontend
UPLOAD_DIR   = BASE_DIR / 'uploads'
# Ensure upload directory exists (create parents if needed)
UPLOAD_DIR.mkdir(parents=True, exist_ok=True)

app = Flask(__name__, static_folder=str(FRONTEND_DIR), static_url_path='')
app.secret_key = 'nmdc-hr-dashboard-secret-key-2024'
CORS(app, supports_credentials=True)

# ─── In-memory store (session-based per user) ──────────────────────────────
# For simplicity, we use a single global store.
# In production, use a database (SQLite/PostgreSQL) per user session.
MASTER_DATA = []   # list of normalized employee dicts
TEMP_FILES = {}    # file_id → {'name': str, 'headers': list, 'rows': list[dict]}

# ─── Constants ───────────────────────────────────────────────────────────────
PROD_DEPTS    = ['Mining','Services (Mech.)','Services (Elect.)','Plant (Mech.)','Plant (Elect.)','Geology & QC','Chemical Lab']
NONPROD_DEPTS = ['Civil','Materials','T&S and Environment','Finance','Human Resource','Industrial Engineering','M&S','C & IT',
                  'Commercial','Contracts Dept.','Works','Vigilance','CSR','ED Sectt.','GM (P) Sectt.','IE Dept.']
OTHER_DEPTS   = ['School','Hospital']
ALL_DEPTS     = PROD_DEPTS + NONPROD_DEPTS + OTHER_DEPTS
DEPOSITS      = ['14','11C','11B']

PROD_SET    = set(PROD_DEPTS)
NONPROD_SET = set(NONPROD_DEPTS)
OTHER_SET   = set(OTHER_DEPTS)

# Column mapping for the specific NMDC test Excel file
NMDC_EXCEL_COLUMN_MAP = {
    'name':        '  N A M E',       # leading spaces — exact as in file
    'department':  'Department',
    'designation': 'Designation',
    'grade':       'GRADE',
    'deposit':     ' DC',              # leading space — exact as in file
    'skills':      'Original Skill',
    'gender':      'Gender',           # added by enrichment script
    'section':     'Section',
    'emp_no':      'UEC No.',
    'sap_no':      'SAP UEC No.',
    'dob':         'DOB',
    'dor':         'DOR',
    'qualification': 'Qualification',
    'category_col': 'Prod./Non-Prod./S&H',
}

# ─── Normalizers ─────────────────────────────────────────────────────────────

def normalize_department(raw):
    if not raw or str(raw).strip() in ('', 'None', 'nan', 'undefined', 'null'):
        return None
    s = str(raw).strip()
    # Pre-aliases: map known Excel variant spellings to canonical dept names
    PRE_ALIAS = {
        'T&S, E':  'T&S and Environment',
        'T&S,E':   'T&S and Environment',
    }
    s = PRE_ALIAS.get(s, s)
    # Exact match (case-insensitive) against known departments
    for d in ALL_DEPTS:
        if d.lower() == s.lower():
            return d
    sl = s.lower()
    # Handle actual department names from the NMDC Excel file
    if 'plant' in sl and ('elec' in sl or 'elect' in sl):  return 'Plant (Elect.)'
    if 'plant' in sl and ('mech' in sl):                    return 'Plant (Mech.)'
    if 'plant' in sl:                                       return 'Plant (Mech.)'
    if 'service' in sl and ('elec' in sl or 'elect' in sl):return 'Services (Elect.)'
    if 'service' in sl:                                     return 'Services (Mech.)'
    if 'mining' in sl:                                      return 'Mining'
    if 'geo' in sl:                                         return 'Geology & QC'
    if 'chem' in sl:                                        return 'Chemical Lab'
    if 'civil' in sl:                                       return 'Civil'
    if 'finance' in sl:                                     return 'Finance'
    if 'human' in sl or re.search(r'\bhr\b', sl):          return 'Human Resource'
    if 'material' in sl:                                    return 'Materials'
    if 'school' in sl:                                      return 'School'
    if 'hospital' in sl:                                    return 'Hospital'
    if 'm&s' in sl or 'm & s' in sl:                       return 'M&S'
    if 'commercial' in sl:                                  return 'Commercial'
    if 'contract' in sl:                                    return 'Contracts Dept.'
    if 'works' in sl:                                       return 'Works'
    if 'vigilance' in sl:                                   return 'Vigilance'
    if 'csr' in sl:                                         return 'CSR'
    if 'ie dept' in sl or 'industrial eng' in sl:          return 'IE Dept.'
    # Matches both 'T&S and Environment' and 'T&S, E' (actual Excel value)
    if 't&s' in sl or 'environment' in sl or re.search(r't&s\s*,', sl): return 'T&S and Environment'
    if 'c &' in sl or ('c' in sl and 'it' in sl):          return 'C & IT'
    return s  # keep unknown as-is


def normalize_deposit(raw):
    if not raw or str(raw).strip() in ('', 'None', 'nan'):
        return None
    s = str(raw).strip().upper()
    # Handle formats: 'Dep-11B', '11B', 'DEP 11C', etc.
    if '11B' in s: return '11B'
    if '11C' in s: return '11C'
    if '-14' in s or s == '14' or s.endswith('14'): return '14'
    return None


def normalize_gender(raw):
    if not raw or str(raw).strip() in ('', 'None', 'nan'):
        return None
    s = str(raw).strip().lower()
    if s in ('m', 'male'):   return 'Male'
    if s in ('f', 'female'): return 'Female'
    return None


# Matches sections like 'Excavation-11B', 'Field Ser-11C(Ele.)', 'CP&DH(E)-14'
# The deposit group may be followed by optional parenthetical suffixes
SECTION_DEPOSIT_RE = re.compile(r'^(?P<base>.*?)[\s\-_/]*(?P<deposit>11B|11C|14)[()\w.\s]*$', re.IGNORECASE)


def normalize_section(raw):
    if not raw or str(raw).strip() in ('', 'None', 'nan', 'undefined', 'null'):
        return None
    return str(raw).strip()


def split_section(raw):
    """Return (base_section, deposit_suffix) for labels like Excavation-11C."""
    section = normalize_section(raw)
    if not section:
        return None, None
    match = SECTION_DEPOSIT_RE.match(section)
    if not match:
        return section, None
    base = match.group('base').strip(' -_/') or section
    return base, match.group('deposit').upper()


def get_category(dept):
    if dept in PROD_SET:    return 'Production'
    if dept in NONPROD_SET: return 'Non Production'
    if dept in OTHER_SET:   return 'Others'
    return 'Unknown'


# Summary workbook loading is done after helper definitions so safe_str is available.


# ─── Virtual department buckets ─────────────────────────────────────────────
# 'C & IT' in the sidebar is a UI bucket that groups all small Non-Prod admin depts
CIT_DEPTS = ['Commercial','Contracts Dept.','Works','Vigilance','CSR','IE Dept.',
             'ED Sectt.','GM (P) Sectt.','M&S','T&S and Environment']

def apply_dept_filter(records, dept_filter):
    """Filter records by department, expanding the 'C & IT' virtual bucket."""
    if dept_filter == 'All':
        return records
    if dept_filter == 'C & IT':
        return [r for r in records if r.get('department') in CIT_DEPTS]
    return [r for r in records if r.get('department') == dept_filter]


# ─── Helpers ─────────────────────────────────────────────────────────────────

def count_by(records, field):
    """Count occurrences of each value in 'field', return sorted list of [value, count] pairs."""
    counter = defaultdict(int)
    for r in records:
        v = r.get(field)
        if v and str(v) not in ('', 'undefined', 'null', 'None'):
            counter[v] += 1
    return sorted(counter.items(), key=lambda x: -x[1])


def safe_str(val, max_len=None):
    s = '' if val is None or (isinstance(val, float) and pd.isna(val)) else str(val).strip()
    if max_len and len(s) > max_len:
        return s[:max_len] + '…'
    return s


def build_grade_pareto_from_summary(summary, dept_filter='All', cat_filter='All'):
    """Build grade_pareto data from REQUIRED_SUMMARY."""
    grade_pareto = []
    total_required = 0
    
    for grade, grade_data in summary['grades'].items():
        # Filter by department if specified
        if dept_filter != 'All':
            dept_info = grade_data.get('departments', {}).get(dept_filter, {})
            count = dept_info.get('required', 0)
        elif cat_filter != 'All':
            # Sum all departments in this category
            count = 0
            for dept, dept_info in grade_data.get('departments', {}).items():
                if get_category(dept) == cat_filter:
                    count += dept_info.get('required', 0)
        else:
            count = grade_data.get('required', 0)
        
        if count > 0:
            total_required += count
            grade_pareto.append({'grade': grade, 'count': count})
    
    # Sort by count descending
    grade_pareto.sort(key=lambda x: -x['count'])
    
    # Calculate cumulative percentages
    cum = 0
    for item in grade_pareto:
        cum += item['count']
        item['cumulative_pct'] = round(cum / total_required * 100) if total_required else 0
    
    return grade_pareto


def build_departments_dist_from_summary(summary, dept_filter='All', cat_filter='All'):
    """Build departments_dist data from REQUIRED_SUMMARY."""
    dept_dist = []
    
    for dept, dept_data in summary['departments'].items():
        # Filter by category if specified
        if cat_filter != 'All' and get_category(dept) != cat_filter:
            continue
        
        count = dept_data.get('required', 0)
        dept_dist.append({'label': dept, 'count': count})
    
    # Sort by count descending
    dept_dist.sort(key=lambda x: -x['count'])
    return dept_dist


SUMMARY_EXCEL_PATH = BASE_DIR.parent / 'Employee_Grade_Dept_Summary.xlsx'
REQUIRED_SUMMARY = None  # Will only be loaded when explicitly uploaded


def parse_summary_workbook(sheet):
    """Parse Employee_Grade_Dept_Summary.xlsx sheet and return summary dict."""
    summary = {
        'enabled': False,
        'total_current': 0,
        'total_required': 0,
        'categories': {
            'Production': {'current': 0, 'required': 0},
            'Non Production': {'current': 0, 'required': 0},
            'Others': {'current': 0, 'required': 0},
            'Unknown': {'current': 0, 'required': 0},
        },
        'departments': {},
        'grades': {},
    }

    try:
        grade_rows = []
        dept_totals = {}

        def parse_int(value):
            try:
                if isinstance(value, (int, float)):
                    return int(value)
                return int(str(value or '').replace(',', '').strip() or '0')
            except (ValueError, TypeError):
                return 0

        for r in range(3, sheet.max_row + 1):
            grade_val = sheet.cell(row=r, column=1).value
            if grade_val is None:
                continue
            grade = safe_str(grade_val).strip()
            if not grade:
                continue

            dept_val = sheet.cell(row=r, column=2).value
            dept_raw = safe_str(dept_val)
            dept = normalize_department(dept_raw) or 'Unknown'

            current = parse_int(sheet.cell(row=r, column=4).value)
            required = parse_int(sheet.cell(row=r, column=5).value)

            if grade.lower() == 'grade' or grade.startswith('▶'):
                continue
            if 'grand total' in grade.lower():
                summary['total_current'] = current
                summary['total_required'] = required
                continue
            if 'subtotal' in grade.lower():
                dept_entry = dept_totals.setdefault(dept, {'current': 0, 'required': 0})
                dept_entry['current'] += current
                dept_entry['required'] += required
                continue

            grade_rows.append({'grade': grade, 'dept': dept, 'current': current})

        # Build departments and categories from subtotal rows.
        for dept, totals in dept_totals.items():
            category = get_category(dept)
            if category not in summary['categories']:
                summary['categories'][category] = {'current': 0, 'required': 0}
            summary['departments'][dept] = {
                'current': totals['current'],
                'required': totals['required']
            }
            summary['categories'][category]['current'] += totals['current']
            summary['categories'][category]['required'] += totals['required']

        # Allocate required counts across grades per department using current proportions.
        dept_grade_current = defaultdict(lambda: defaultdict(int))
        grade_current = defaultdict(int)
        for row in grade_rows:
            dept_grade_current[row['dept']][row['grade']] += row['current']
            grade_current[row['grade']] += row['current']

        def allocate_counts(weights, total):
            if total <= 0 or not weights:
                return {k: 0 for k in weights}
            total_weight = sum(weights.values())
            if total_weight <= 0:
                per = total // len(weights)
                return {k: per for k in weights}

            alloc = {}
            remainders = []
            allocated = 0
            for key, weight in weights.items():
                exact = total * weight / total_weight
                floor_val = int(exact)
                alloc[key] = floor_val
                allocated += floor_val
                remainders.append((key, exact - floor_val))

            remainder = total - allocated
            remainders.sort(key=lambda x: -x[1])
            for key, _ in remainders[:remainder]:
                alloc[key] += 1
            return alloc

        grade_required_by_dept = defaultdict(dict)
        for dept, grade_counts in dept_grade_current.items():
            required_total = dept_totals.get(dept, {}).get('required', 0)
            allocations = allocate_counts(grade_counts, required_total)
            for grade, alloc in allocations.items():
                grade_required_by_dept[grade][dept] = alloc

        # Build grade summary using current counts and allocated required counts.
        for grade, current in grade_current.items():
            summary['grades'][grade] = {'current': current, 'required': 0, 'departments': {}}

        for grade, dept_allocs in grade_required_by_dept.items():
            for dept, alloc in dept_allocs.items():
                grade_entry = summary['grades'][grade]
                grade_entry['required'] += alloc
                grade_entry['departments'][dept] = {
                    'current': dept_grade_current[dept][grade],
                    'required': alloc
                }

        # If the grand total row was unavailable or formula values were not present,
        # derive totals from subtotal department values as a fallback.
        if summary['total_current'] == 0 and summary['total_required'] == 0:
            summary['total_current'] = sum(v['current'] for v in summary['departments'].values())
            summary['total_required'] = sum(v['required'] for v in summary['departments'].values())

        summary['enabled'] = True
    except Exception as e:
        app.logger.error('Error parsing summary workbook: %s', e)
        summary['enabled'] = False

    return summary


def load_required_summary():
    """Legacy function — loads from disk if file exists."""
    if not SUMMARY_EXCEL_PATH.exists():
        return None
    try:
        from openpyxl import load_workbook
        wb = load_workbook(SUMMARY_EXCEL_PATH, data_only=True)
        return parse_summary_workbook(wb.active)
    except Exception as exc:
        app.logger.error('Failed to load summary workbook %s: %s', SUMMARY_EXCEL_PATH, exc)
        return None


# Don't auto-load summary file — require explicit user upload
# REQUIRED_SUMMARY = load_required_summary()


def parse_skills(skills_str):
    """Split a skills string into individual skill tokens."""
    if not skills_str or str(skills_str).strip() in ('', 'nan', 'None'):
        return []
    return [t.strip() for t in re.split(r'[,;|/]', str(skills_str)) if t.strip() and t.strip() not in ('undefined','null','None')]


# ─── Routes ──────────────────────────────────────────────────────────────────

@app.route('/')
def index():
    """Serve the frontend index.html."""
    return send_file(str(FRONTEND_DIR / 'index.html'))


@app.route('/api/autoload', methods=['POST'])
def autoload_excel():
    """
    Auto-load Excel file. Detects if it's Employee_Grade_Dept_Summary or regular employee data.
    For summary file: parses and sets REQUIRED_SUMMARY.
    For regular files: auto-loads with pre-known column mapping.
    Returns: count of records loaded and summary status.
    """
    global MASTER_DATA, REQUIRED_SUMMARY

    if 'file' not in request.files:
        return jsonify({'error': 'No file provided'}), 400

    f = request.files['file']
    if not f.filename:
        return jsonify({'error': 'Empty filename'}), 400

    file_bytes = f.read()
    filename = f.filename.lower()

    # Check if this is the summary file
    if 'grade_dept_summary' in filename or 'employee_strength' in filename:
        try:
            from openpyxl import load_workbook
            wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
            sheet = wb.active
            REQUIRED_SUMMARY = parse_summary_workbook(sheet)
            if REQUIRED_SUMMARY and REQUIRED_SUMMARY.get('enabled'):
                return jsonify({
                    'status': 'summary_loaded',
                    'message': f'Summary file loaded: {REQUIRED_SUMMARY["total_current"]}/{REQUIRED_SUMMARY["total_required"]} employees',
                    'loaded': 0,
                    'total': len(MASTER_DATA),
                    'summary_enabled': True
                })
            else:
                return jsonify({'error': 'Failed to parse summary file'}), 422
        except Exception as e:
            app.logger.error('Failed to process summary file: %s', e)
            return jsonify({'error': f'Failed to process summary file: {str(e)}'}), 422

    # Regular employee data processing
    try:
        if filename.endswith('.csv'):
            df = pd.read_csv(io.BytesIO(file_bytes), dtype=str, na_filter=False)
        else:
            df = pd.read_excel(io.BytesIO(file_bytes), dtype=str, na_filter=False)
    except Exception as e:
        return jsonify({'error': f'Failed to parse file: {str(e)}'}), 422

    if df.empty:
        return jsonify({'error': 'File contains no data'}), 422

    # Clean up column names and values
    df.columns = [str(c) for c in df.columns]  # keep exact names with spaces
    df = df.replace({'nan': '', 'NaN': '', '<NA>': '', 'None': ''})

    # Detect which columns are available in this file
    col_map = NMDC_EXCEL_COLUMN_MAP
    available = set(df.columns)

    def get_col(key):
        col = col_map.get(key)
        return col if col and col in available else None

    dept_col  = get_col('department')
    name_col  = get_col('name')
    desig_col = get_col('designation')
    grade_col = get_col('grade')
    dep_col   = get_col('deposit')
    skill_col = get_col('skills')
    gender_col = get_col('gender')

    if not dept_col:
        return jsonify({'error': 'Department column not found in file. Columns: ' + str(list(df.columns))}), 422

    records = []
    skipped = 0
    for _, row in df.iterrows():
        dept_raw = safe_str(row.get(dept_col, ''))
        dept = normalize_department(dept_raw)
        if not dept:
            skipped += 1
            continue

        skills_raw  = safe_str(row.get(skill_col, '')) if skill_col else ''
        grade_raw   = safe_str(row.get(grade_col, '')).strip() if grade_col else ''  # strip trailing spaces from grades like 'RS1 '
        deposit_raw = safe_str(row.get(dep_col, '')) if dep_col else ''
        gender_raw  = safe_str(row.get(gender_col, '')) if gender_col else ''

        records.append({
            'name':        safe_str(row.get(name_col, '')) if name_col else '',
            'department':  dept,
            'designation': safe_str(row.get(desig_col, '')) if desig_col else '',
            'grade':       grade_raw,
            'deposit':     normalize_deposit(deposit_raw),
            'gender':      normalize_gender(gender_raw),
            'skills':      skills_raw,
            'skills_list': parse_skills(skills_raw),
            'category':    get_category(dept),
            '_file':       f.filename,
            # Extra columns from NMDC file
            'emp_no':      safe_str(row.get(col_map.get('emp_no',''), '')) if col_map.get('emp_no') in available else '',
            'section':     safe_str(row.get(col_map.get('section',''), '')) if col_map.get('section') in available else '',
            'qualification': safe_str(row.get(col_map.get('qualification',''), '')) if col_map.get('qualification') in available else '',
        })

    MASTER_DATA.extend(records)

    unique_depts = len(set(r['department'] for r in MASTER_DATA))
    return jsonify({
        'loaded':       len(records),
        'skipped':      skipped,
        'total':        len(MASTER_DATA),
        'unique_depts': unique_depts,
        'summary_enabled': REQUIRED_SUMMARY is not None and REQUIRED_SUMMARY.get('enabled', False),
        'columns_used': {
            'department':  dept_col,
            'name':        name_col,
            'designation': desig_col,
            'grade':       grade_col,
            'deposit':     dep_col,
            'skills':      skill_col,
        }
    })


@app.route('/api/upload', methods=['POST'])
def upload_file():
    """
    Upload an Excel (.xlsx/.xls) or CSV file.
    Returns:
        file_id, file_name, headers (column names), preview (first 3 rows as list of dicts)
    """
    if 'file' not in request.files:
        return jsonify({'error': 'No file provided'}), 400

    f = request.files['file']
    if not f.filename:
        return jsonify({'error': 'Empty filename'}), 400

    filename = f.filename.lower()
    file_bytes = f.read()

    try:
        if filename.endswith('.csv'):
            df = pd.read_csv(io.BytesIO(file_bytes), dtype=str, na_filter=False)
        else:
            df = pd.read_excel(io.BytesIO(file_bytes), dtype=str, na_filter=False)
    except Exception as e:
        return jsonify({'error': f'Failed to parse file: {str(e)}'}), 422

    if df.empty:
        return jsonify({'error': 'File contains no data'}), 422

    # Replace NaN strings
    df = df.replace({'nan': '', 'NaN': '', '<NA>': '', 'None': ''})
    df.columns = [str(c).strip() for c in df.columns]

    headers = df.columns.tolist()
    rows = df.to_dict(orient='records')

    # Preview — first 3 rows, first 8 columns
    preview_cols = headers[:8]
    preview = []
    for i, row in enumerate(rows[:3]):
        preview.append({c: safe_str(row.get(c, ''), 30) for c in preview_cols})

    file_id = str(uuid.uuid4())
    TEMP_FILES[file_id] = {
        'name': f.filename,
        'headers': headers,
        'rows': rows,
        'bytes': file_bytes
    }

    return jsonify({
        'file_id': file_id,
        'file_name': f.filename,
        'headers': headers,
        'row_count': len(rows),
        'preview': preview,
        'preview_cols': preview_cols
    })


@app.route('/api/process', methods=['POST'])
def process_file():
    """
    Apply column mapping to a previously uploaded file.
    If the file is Employee_Grade_Dept_Summary, load it as REQUIRED_SUMMARY.
    Otherwise, parse as employee data and add to MASTER_DATA.
    Body JSON:
        file_id: str
        mapping: {department, name, designation, grade, deposit, gender, skills}
                 Values are column names (strings). (optional for summary files)
    Returns: count of records added, total master count.
    """
    global MASTER_DATA, REQUIRED_SUMMARY

    body = request.get_json(force=True)
    file_id = body.get('file_id')
    mapping = body.get('mapping', {})

    if not file_id or file_id not in TEMP_FILES:
        return jsonify({'error': 'Unknown file_id'}), 400

    file_item = TEMP_FILES[file_id]
    file_name = file_item['name'].lower()

    # Check if this is the summary file
    if 'grade_dept_summary' in file_name or 'employee_strength' in file_name:
        try:
            from openpyxl import load_workbook
            file_bytes = io.BytesIO(file_item.get('bytes', b''))
            wb = load_workbook(file_bytes, data_only=True)
            sheet = wb.active
            # Parse summary and store in REQUIRED_SUMMARY
            REQUIRED_SUMMARY = parse_summary_workbook(sheet)
            # Clean up temp file
            del TEMP_FILES[file_id]
            if REQUIRED_SUMMARY and REQUIRED_SUMMARY.get('enabled'):
                return jsonify({
                    'status': 'summary_loaded',
                    'message': f'Summary file loaded: {REQUIRED_SUMMARY["total_current"]}/{REQUIRED_SUMMARY["total_required"]} employees',
                    'total': len(MASTER_DATA),
                    'summary_enabled': True
                })
            else:
                return jsonify({'error': 'Failed to parse summary file'}), 422
        except Exception as e:
            app.logger.error('Failed to process summary file: %s', e)
            del TEMP_FILES[file_id]
            return jsonify({'error': f'Failed to process summary file: {str(e)}'}), 422

    # Regular employee data processing
    dept_col = mapping.get('department')
    if not dept_col:
        return jsonify({'error': 'Department column mapping is required'}), 400

    rows = file_item['rows']

    # Validate mapping columns exist in uploaded file headers
    headers = file_item.get('headers', [])
    missing_cols = [m for m in set(mapping.values()) if m and m not in headers]
    if missing_cols:
        return jsonify({'error': 'Mapping refers to unknown columns: ' + ', '.join(missing_cols)}), 400

    records = []
    for row in rows:
        dept_raw = row.get(dept_col, '')
        dept = normalize_department(dept_raw)
        if not dept:
            continue

        def get_val(col_name):
            if not col_name:
                return ''
            return safe_str(row.get(col_name, ''))

        skills_raw = get_val(mapping.get('skills'))
        records.append({
            'name':        get_val(mapping.get('name')),
            'department':  dept,
            'designation': get_val(mapping.get('designation')),
            'grade':       get_val(mapping.get('grade')),
            'deposit':     normalize_deposit(get_val(mapping.get('deposit'))),
            'gender':      normalize_gender(get_val(mapping.get('gender'))),
            'skills':      skills_raw,
            'skills_list': parse_skills(skills_raw),
            'category':    get_category(dept),
            '_file':       file_name,
        })

    MASTER_DATA.extend(records)

    # Clean up temp file from memory
    del TEMP_FILES[file_id]

    return jsonify({
        'added': len(records),
        'total': len(MASTER_DATA),
        'unique_depts': len(set(r['department'] for r in MASTER_DATA)),
        'summary_enabled': REQUIRED_SUMMARY is not None and REQUIRED_SUMMARY.get('enabled', False)
    })



@app.route('/api/analytics', methods=['GET'])
def analytics():
    """
    Return aggregated analytics for the current data with optional filters.
    Query params:
        deposit: All | 11B | 11C | 14
        cat:     All | Production | Non Production | Others
        dept:    All | <department name>
        show:    Required | (leave empty for regular data)
    """
    deposit_filter = request.args.get('deposit', 'All')
    cat_filter     = request.args.get('cat', 'All')
    dept_filter    = request.args.get('dept', 'All')
    show_mode      = request.args.get('show', '')

    df = list(MASTER_DATA)

    # Apply filters
    if deposit_filter != 'All':
        df = [r for r in df if r.get('deposit') == deposit_filter]
    if cat_filter != 'All':
        df = [r for r in df if r.get('category') == cat_filter]
    df = apply_dept_filter(df, dept_filter)

    n = len(df)

    # ── KPIs ────────────────────────────────────────────────────────────────
    grades      = set(r['grade'] for r in df if r.get('grade'))
    desigs      = set(r['designation'] for r in df if r.get('designation'))
    depts_set   = set(r['department'] for r in df if r.get('department'))
    female_cnt  = sum(1 for r in df if r.get('gender') == 'Female')
    female_pct  = round(female_cnt / n * 100) if n else 0

    # ── Grade distribution (Pareto) ──────────────────────────────────────
    # Use summary data if show_mode='Required'
    if show_mode == 'Required' and REQUIRED_SUMMARY and REQUIRED_SUMMARY.get('enabled'):
        grade_pareto = build_grade_pareto_from_summary(REQUIRED_SUMMARY, dept_filter, cat_filter)
    else:
        grade_counts = count_by(df, 'grade')
        total_grades = sum(v for _, v in grade_counts)
        cum = 0
        grade_pareto = []
        for label, val in grade_counts:
            cum += val
            grade_pareto.append({
                'grade': label,
                'count': val,
                'cumulative_pct': round(cum / total_grades * 100) if total_grades else 0
            })

    # ── Gender ──────────────────────────────────────────────────────────────
    gender_counts = count_by(df, 'gender')

    # ── Deposit distribution ─────────────────────────────────────────────
    deposit_counts = count_by(df, 'deposit')

    # ── Deposit performance cards ────────────────────────────────────────
    # Show only the selected deposit card when filtered, else show all
    deps_to_show = [deposit_filter] if deposit_filter != 'All' else DEPOSITS
    deposit_cards = []
    for dep in deps_to_show:
        sub      = [r for r in df if r.get('deposit') == dep]
        all_sub  = [r for r in MASTER_DATA if r.get('deposit') == dep]
        emp      = len(sub)
        total_emp = len(all_sub)
        fem_cnt  = sum(1 for r in sub if r.get('gender') == 'Female')
        fem_pct  = round(fem_cnt / emp * 100) if emp else 0
        g_set    = set(r['grade'] for r in sub if r.get('grade'))
        d_set    = set(r['designation'] for r in sub if r.get('designation'))
        sk_set   = set()
        for r in sub:
            sk_set.update(r.get('skills_list', []))
        pct = round(emp / n * 100) if n else 0
        emp_bar_pct = round(emp / total_emp * 100) if total_emp else 0
        deposit_cards.append({
            'deposit':     dep,
            'employees':   emp,
            'total_employees': total_emp,
            'female_pct':  fem_pct,
            'grade_count': len(g_set),
            'desig_count': len(d_set),
            'skill_count': len(sk_set),
            'pct':         pct,
            'emp_bar_pct': emp_bar_pct
        })

    # ── Top designations ─────────────────────────────────────────────────
    desig_counts_all = count_by(df, 'designation')
    desig_counts = desig_counts_all[:10]

    # ── Skills treemap ───────────────────────────────────────────────────
    skill_counter = defaultdict(int)
    for r in df:
        for sk in r.get('skills_list', []):
            if sk and sk not in ('undefined', 'null'):
                skill_counter[sk] += 1
    skills_data = sorted(skill_counter.items(), key=lambda x: -x[1])[:20]

    # ── Workforce summary panel ──────────────────────────────────────────
    top_dept   = desig_counts  # reuse count_by for dept below
    dept_counts = count_by(df, 'department')
    top_dept_name  = dept_counts[0][0] if dept_counts else '—'
    top_dep_name   = deposit_counts[0][0] if deposit_counts else '—'
    top_skill_name = skills_data[0][0] if skills_data else '—'
    top_desig_name = desig_counts[0][0] if desig_counts else '—'

    # ── Insights cards ────────────────────────────────────────────────────
    avg_per_dept = round(n / len(dept_counts)) if dept_counts else 0
    top_grade    = grade_pareto[0] if grade_pareto else None
    top_deposit  = deposit_counts[0] if deposit_counts else None

    # ── Sidebar counts ────────────────────────────────────────────────────
    base = MASTER_DATA if deposit_filter == 'All' else [r for r in MASTER_DATA if r.get('deposit') == deposit_filter]
    dept_count_map = defaultdict(int)
    for r in base:
        dept_count_map[r['department']] += 1

    summary_info = None
    if show_mode == 'Required' and REQUIRED_SUMMARY and REQUIRED_SUMMARY.get('enabled'):
        summary_info = {
            'mode': 'Required',
            'total_current': REQUIRED_SUMMARY['total_current'],
            'total_required': REQUIRED_SUMMARY['total_required'],
            'categories': {
                cat: {'current': vals['current'], 'required': vals['required']}
                for cat, vals in REQUIRED_SUMMARY['categories'].items()
            },
            'departments': REQUIRED_SUMMARY['departments'],
            'grades': REQUIRED_SUMMARY['grades']
        }

    sidebar = {
        'all':      len(base),
        'prod_all': sum(dept_count_map[d] for d in PROD_DEPTS),
        'np_all':   sum(dept_count_map[d] for d in NONPROD_DEPTS),
        'oth_all':  sum(dept_count_map[d] for d in OTHER_DEPTS),
        'by_dept':  dict(dept_count_map)
    }

    return jsonify({
        'total': n,
        'kpis': {
            'total':      n,
            'departments': len(depts_set),
            'grades':     len(grades),
            'female_pct': female_pct,
            'designations': len(desigs)
        },
        'grade_pareto':   [{'label': g['grade'], 'count': g['count'], 'cum_pct': g['cumulative_pct']} for g in grade_pareto],
        'gender':         [{'label': l, 'count': c} for l, c in gender_counts],
        'deposit_dist':   [{'label': l, 'count': c} for l, c in deposit_counts],
        'deposit_cards':  deposit_cards,
        'departments_dist': [{'label': d['label'], 'count': d['count']} for d in (build_departments_dist_from_summary(REQUIRED_SUMMARY, dept_filter, cat_filter) if show_mode == 'Required' and REQUIRED_SUMMARY and REQUIRED_SUMMARY.get('enabled') else [{'label': l, 'count': c} for l, c in dept_counts])],
        'designations':   [{'label': l, 'count': c} for l, c in desig_counts],
        'designations_dist': [{'label': l, 'count': c} for l, c in desig_counts_all],
        'skills':         [{'label': l, 'count': c} for l, c in skills_data],
        'workforce_summary': {
            'total':       n,
            'top_deposit': top_dep_name,
            'top_dept':    top_dept_name,
            'top_skill':   top_skill_name,
            'top_desig':   top_desig_name,
            'female_pct':  female_pct
        },
        'insights': {
            'avg_per_dept':   avg_per_dept,
            'top_grade':      {'label': top_grade['grade'], 'count': top_grade['count']} if top_grade else None,
            'largest_dept':   {'label': dept_counts[0][0], 'count': dept_counts[0][1]} if dept_counts else None,
            'largest_deposit': {'label': top_deposit[0], 'count': top_deposit[1]} if top_deposit else None,
            'top_desig':      {'label': desig_counts[0][0], 'count': desig_counts[0][1]} if desig_counts else None,
            'top_skill':      {'label': skills_data[0][0], 'count': skills_data[0][1]} if skills_data else None,
            'dept_count':     len(dept_counts)
        },
        'sidebar': sidebar,
        'summary': summary_info
    })


@app.route('/api/sections', methods=['GET'])
def sections():
    """
    Return section breakdown split by deposit.
    Query params:
        dept:    specific department name, or omit for category-level
        cat:     category filter (Production | Non Production | Others) — used when dept not set
        deposit: All | 11B | 11C | 14
    """
    dept_filter    = request.args.get('dept', 'All')
    cat_filter     = request.args.get('cat', 'All')
    deposit_filter = request.args.get('deposit', 'All')

    df = list(MASTER_DATA)

    # Apply department filter (expands 'C & IT' virtual bucket)
    df = apply_dept_filter(df, dept_filter)
    if dept_filter == 'All' and cat_filter != 'All':
        df = [r for r in df if r.get('category') == cat_filter]

    # Apply deposit filter
    if deposit_filter != 'All':
        df = [r for r in df if r.get('deposit') == deposit_filter]

    # Build section × deposit matrix, grouped by department + base section.
    section_map = defaultdict(lambda: defaultdict(int))
    for r in df:
        sec_base, _ = split_section(r.get('section', ''))
        sec = sec_base or '(No Section)'
        dep = r.get('deposit') or 'Unknown'
        dept = r.get('department', '')
        section_map[(dept, sec)][dep] += 1

    rows = []
    for (dept, sec), dep_counts in section_map.items():
        total = sum(dep_counts.values())
        rows.append({
            'section': sec,
            'dept':    dept,
            'counts':  {d: dep_counts.get(d, 0) for d in DEPOSITS},
            'unknown': dep_counts.get('Unknown', 0),
            'total':   total
        })
    rows.sort(key=lambda x: (x['dept'] or '', -x['total'], x['section']))
    # Determine which deposits are actually present in the filtered dataset.
    deposits_present = [d for d in DEPOSITS if any(r.get('deposit') == d for r in df)]

    # If the request is for Non Production or Others (category-level), or the
    # department requested belongs to those categories, we intentionally hide
    # the deposit columns and only return section-wise totals as requested.
    hide_deposits = False
    if cat_filter in ('Non Production', 'Others'):
        hide_deposits = True
    if dept_filter == 'C & IT':
        hide_deposits = True
    if dept_filter != 'All' and dept_filter not in ('C & IT',) and (dept_filter in NONPROD_SET or dept_filter in OTHER_SET):
        hide_deposits = True

    if hide_deposits:
        deposits_present = []
        # remove per-deposit breakdown from rows to simplify frontend rendering
        for r in rows:
            r['counts'] = {}

    return jsonify({
        'dept':     dept_filter,
        'cat':      cat_filter,
        'deposit':  deposit_filter,
        'deposits': deposits_present,
        'rows':     rows,
        'total':    len(df)
    })


@app.route('/api/section-details', methods=['GET'])
def section_details():
    """
    Return employee-level rows for a specific department + section + deposit.
    Query params:
        dept:    department name, or All when filtering by category
        cat:     category filter when dept is All
        deposit: All | 11B | 11C | 14
        section: base section name from /api/sections
    """
    dept_filter    = request.args.get('dept', 'All')
    cat_filter     = request.args.get('cat', 'All')
    deposit_filter = request.args.get('deposit', 'All')
    section_filter = request.args.get('section', 'All')

    df = list(MASTER_DATA)

    # Apply department filter (expands 'C & IT' virtual bucket)
    df = apply_dept_filter(df, dept_filter)
    if dept_filter == 'All' and cat_filter != 'All':
        df = [r for r in df if r.get('category') == cat_filter]

    if deposit_filter != 'All':
        df = [r for r in df if r.get('deposit') == deposit_filter]

    section_key = (section_filter or '').strip().lower()
    if section_key != 'all':
        filtered = []
        for r in df:
            base, _ = split_section(r.get('section', ''))
            if (base or '').strip().lower() == section_key:
                filtered.append(r)
        df = filtered

    rows = []
    for r in df:
        base_section, section_deposit = split_section(r.get('section', ''))
        rows.append({
            'name': r.get('name', ''),
            'department': r.get('department', ''),
            'section': r.get('section', ''),
            'base_section': base_section or '',
            'section_deposit': section_deposit or '',
            'deposit': r.get('deposit', ''),
            'designation': r.get('designation', ''),
            'grade': r.get('grade', ''),
            'gender': r.get('gender', ''),
            'skills': ', '.join(r.get('skills_list', [])) if r.get('skills_list') else r.get('skills', ''),
            'emp_no': r.get('emp_no', ''),
            'qualification': r.get('qualification', '')
        })

    rows.sort(key=lambda x: (x['department'], x['section'], x['name']))

    return jsonify({
        'dept': dept_filter,
        'cat': cat_filter,
        'deposit': deposit_filter,
        'section': section_filter,
        'total': len(rows),
        'rows': rows
    })


@app.route('/api/departments', methods=['GET'])
def departments():
    """
    Return department list with employee counts for a given category.
    Query params:
        category: Production | Non Production | Others
    """
    category_filter = request.args.get('category', 'All')
    
    df = list(MASTER_DATA)
    
    # Filter by category
    if category_filter == 'Production':
        df = [r for r in df if r.get('category') == 'Production']
    elif category_filter == 'Non Production':
        df = [r for r in df if r.get('category') == 'Non Production']
    elif category_filter == 'Others':
        df = [r for r in df if r.get('category') == 'Others']
    
    # Count employees by department
    dept_counts = defaultdict(int)
    for r in df:
        dept = r.get('department', 'Unknown')
        dept_counts[dept] += 1
    
    # Build rows sorted by employee count (descending) then department name
    rows = [
        {'department': dept, 'employees': count}
        for dept, count in sorted(dept_counts.items(), key=lambda x: (-x[1], x[0]))
    ]
    
    return jsonify({
        'category': category_filter,
        'rows': rows,
        'total': len(df)
    })

@app.route('/api/summary', methods=['GET'])
def summary():
    """Quick summary for status pill."""
    return jsonify({
        'total': len(MASTER_DATA),
        'files': len(set(r['_file'] for r in MASTER_DATA)) if MASTER_DATA else 0,
        'summary_enabled': REQUIRED_SUMMARY is not None and REQUIRED_SUMMARY.get('enabled', False)
    })


@app.route('/api/employees-by-attribute', methods=['GET'])
def employees_by_attribute():
    """
    Return employee rows for a specific deposit or skill value,
    respecting the current cat / dept filters.
    Query params:
        type:   'deposit' | 'skill'
        value:  the deposit label (11B/11C/14) or skill name
        cat:    All | Production | Non Production | Others
        dept:   All | <department name>
    """
    attr_type   = request.args.get('type', 'deposit')
    attr_value  = request.args.get('value', '')
    cat_filter  = request.args.get('cat', 'All')
    dept_filter = request.args.get('dept', 'All')
    section_filter = request.args.get('section', 'All')

    df = list(MASTER_DATA)

    # Apply category / department filters
    if cat_filter != 'All':
        df = [r for r in df if r.get('category') == cat_filter]
    df = apply_dept_filter(df, dept_filter)

    # Filter by attribute (skip filter if value is 'All' or empty — returns all employees for cat/dept)
    if attr_value and attr_value.lower() != 'all':
        if attr_type == 'deposit':
            df = [r for r in df if r.get('deposit') == attr_value]
        elif attr_type == 'skill':
            av = str(attr_value or '').strip().lower()
            df = [r for r in df if any(av == (s or '').strip().lower() for s in r.get('skills_list', []))]
            # Apply section filter for skills
            if section_filter and section_filter.lower() != 'all':
                df = [r for r in df if r.get('section', '').strip() == section_filter]
        elif attr_type == 'gender':
            df = [r for r in df if r.get('gender') == attr_value]
        elif attr_type == 'grade':
            df = [r for r in df if r.get('grade') == attr_value]
        elif attr_type == 'department':
            df = [r for r in df if r.get('department') == attr_value]
        elif attr_type == 'desig':
            df = [r for r in df if r.get('designation') == attr_value]

    rows = []
    for r in df:
        rows.append({
            'name':        r.get('name', '—'),
            'department':  r.get('department', '—'),
            'designation': r.get('designation', '—'),
            'grade':       r.get('grade', '—'),
            'deposit':     r.get('deposit', '—'),
            'gender':      r.get('gender', '—'),
            'skills':      ', '.join(r.get('skills_list', [])) if r.get('skills_list') else r.get('skills', '—'),
            'emp_no':      r.get('emp_no', '—'),
            'section':     r.get('section', '—'),
        })

    rows.sort(key=lambda x: (x['department'], x['name']))

    return jsonify({
        'type':  attr_type,
        'value': attr_value,
        'total': len(rows),
        'rows':  rows
    })


@app.route('/api/skill-sections', methods=['GET'])
def skill_sections():
    """
    Return unique sections for a given skill.
    Query params:
        skill:  skill name
        cat:    All | Production | Non Production | Others
        dept:   All | <department name>
    """
    skill_val   = request.args.get('skill', '')
    cat_filter  = request.args.get('cat', 'All')
    dept_filter = request.args.get('dept', 'All')

    if not skill_val:
        return jsonify({'error': 'Skill parameter required'}), 400

    df = list(MASTER_DATA)

    # Apply category / department filters
    if cat_filter != 'All':
        df = [r for r in df if r.get('category') == cat_filter]
    df = apply_dept_filter(df, dept_filter)

    # Filter employees that have this skill
    skill_val_lower = str(skill_val or '').strip().lower()
    df = [r for r in df if any(skill_val_lower == (s or '').strip().lower() for s in r.get('skills_list', []))]

    # Get unique sections for these employees
    sections = sorted(set(r.get('section', '').strip() for r in df if r.get('section', '').strip()))

    return jsonify({
        'skill': skill_val,
        'sections': sections,
        'total_employees': len(df),
        'employees_by_section': {
            sec: len([r for r in df if r.get('section', '').strip() == sec])
            for sec in sections
        }
    })


@app.route('/api/clear', methods=['DELETE'])
def clear_data():
    """Clear all server-side data."""
    global MASTER_DATA
    MASTER_DATA = []
    TEMP_FILES.clear()
    return jsonify({'status': 'cleared'})


# ─── Run ─────────────────────────────────────────────────────────────────────
if __name__ == '__main__':
    print("\n" + "=" * 58)
    print("  NMDC HR Analytics Intelligence - Flask Backend")
    print("=" * 58)
    print(f"  Frontend : http://localhost:5000")
    print(f"  API Docs :")
    print(f"    POST   /api/upload    -> Upload Excel/CSV file")
    print(f"    POST   /api/process   -> Apply column mapping")
    print(f"    GET    /api/analytics -> Get aggregated analytics")
    print(f"    DELETE /api/clear     -> Clear all data")
    print("=" * 58 + "\n")
    app.run(debug=True, port=5000, host='0.0.0.0')
