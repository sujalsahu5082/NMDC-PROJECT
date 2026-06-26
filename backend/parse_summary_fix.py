import openpyxl
from pathlib import Path
from collections import defaultdict

wb=openpyxl.load_workbook(Path('Employee_Grade_Dept_Summary.xlsx'), data_only=True)
ws=wb.active

def text(v):
    return str(v).strip() if v is not None else ''

def normalize_department(raw):
    if not raw or str(raw).strip() in ('', 'None', 'nan', 'undefined', 'null'):
        return None
    s=str(raw).strip()
    PRE_ALIAS={'T&S, E':'T&S and Environment','T&S,E':'T&S and Environment'}
    s=PRE_ALIAS.get(s,s)
    if s.lower() == 'sp-iii/works':
        return 'SP-III/Works'
    return s

summary={'total_current':0,'total_required':0,'departments':defaultdict(lambda:{'current':0,'required':0}),'grades':defaultdict(lambda:{'current':0,'required':0,'departments':defaultdict(lambda:{'current':0,'required':0})})}

for r in range(3,229):
    grade_raw=ws.cell(r,1).value
    if grade_raw is None: continue
    grade=text(grade_raw)
    if not grade: continue
    dept_raw=ws.cell(r,2).value
    dept=text(dept_raw)
    cur=ws.cell(r,4).value
    req=ws.cell(r,5).value
    if isinstance(cur,(int,float)):
        cur=int(cur)
    else:
        continue
    if not isinstance(req,(int,float)):
        continue
    req=int(req)
    if grade.lower() == 'grade' or grade.startswith('▶'):
        continue
    if 'grand total' in grade.lower():
        print('grand total row', cur, req)
        summary['total_current']=cur
        summary['total_required']=req
        continue
    if 'subtotal' in grade.lower():
        dept=normalize_department(dept_raw) or 'Unknown'
        summary['departments'][dept]['current']=cur
        summary['departments'][dept]['required']=req
        continue
    summary['grades'][grade]['current'] += cur
    summary['grades'][grade]['required'] += req
    grade_dept = summary['grades'][grade]['departments'][dept]
    grade_dept['current'] += cur
    grade_dept['required'] += req

print('total_current', summary['total_current'])
print('total_required', summary['total_required'])
print('dept required total', sum(d['required'] for d in summary['departments'].values()))
print('grade required total', sum(g['required'] for g in summary['grades'].values()))
print('departments', len(summary['departments']))
print('grades', len(summary['grades']))
print('dept sample', list(summary['departments'].items())[:10])
print('grade sample', {g:summary['grades'][g]['required'] for g in list(summary['grades'])[:10]})
